import asyncio
from playwright.async_api import async_playwright
import os
import time
import sys
import csv
import json
import re
import shutil
import socket
import subprocess
from datetime import datetime
import webbrowser
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog
from time import perf_counter

def log_progress(msg: str):
    """간단 진행 로그 (시간표기/레벨 옵션 제거)"""
    print(msg)

def extract_red_background_cells(excel_file_path):
    """엑셀 파일에서 빨간색 배경을 가진 셀들의 값을 추출"""
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        red_cell_values = []
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # 셀의 배경색 확인
                        fill = cell.fill
                        if fill.fill_type == 'solid':
                            # RGB 값으로 빨간색 계열인지 확인
                            if fill.start_color.rgb:
                                color = fill.start_color.rgb
                                if isinstance(color, str) and len(color) == 8:
                                    # ARGB 형식에서 RGB 추출
                                    r = int(color[2:4], 16)
                                    g = int(color[4:6], 16)
                                    b = int(color[6:8], 16)
                                    
                                    # 빨간색 계열 판단 (빨간색이 가장 강한 색상)
                                    if r > g and r > b and r > 150:
                                        cell_value = str(cell.value).strip()
                                        if cell_value and cell_value not in red_cell_values:
                                            red_cell_values.append(cell_value)
        
        workbook.close()
        print(f"📊 엑셀 파일에서 {len(red_cell_values)}개의 키워드를 추출했습니다.")
        for i, keyword in enumerate(red_cell_values, 1):
            print(f"   ▸ {keyword}")
        
        return red_cell_values
        
    except Exception as e:
        print(f"⚠️ 엑셀 파일 읽기 오류: {e}")
        return []

def select_excel_file():
    """엑셀 파일 선택 대화상자"""
    try:
        # tkinter 창을 숨김
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        file_path = filedialog.askopenfilename(
            title="매칭할 키워드가 있는 엑셀 파일을 선택하세요",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        root.destroy()
        return file_path
        
    except Exception as e:
        print(f"⚠️ 파일 선택 오류: {e}")
        return None

def highlight_keywords_in_text(text, keywords):
    """텍스트에서 키워드와 부분 일치하는 부분을 하이라이트 (중복 하이라이트 방지)"""
    if not text or not keywords:
        return text
    
    # 키워드 길이 순으로 정렬 (긴 키워드부터 처리)
    sorted_keywords = sorted(keywords, key=len, reverse=True)
    
    # 각 키워드에 대해 순차적으로 하이라이트 적용
    result_text = text
    
    for keyword in sorted_keywords:
        if keyword.strip():
            # 이미 하이라이트된 부분을 임시로 보호
            highlight_pattern = r'<span class="highlight">[^<]*</span>'
            protected_parts = []
            temp_text = result_text
            
            # 기존 하이라이트 부분을 임시 플레이스홀더로 대체
            import re
            matches = list(re.finditer(highlight_pattern, temp_text))
            for i, match in enumerate(reversed(matches)):
                placeholder = f"__PROTECTED_{i}__"
                protected_parts.insert(0, match.group())
                temp_text = temp_text[:match.start()] + placeholder + temp_text[match.end():]
            
            # 보호된 텍스트에서 키워드 하이라이트 적용 (단어 경계 고려하지 않음 - 부분 일치)
            pattern = re.escape(keyword.strip())
            temp_text = re.sub(
                f'({pattern})', 
                r'<span class="highlight">\1</span>', 
                temp_text, 
                flags=re.IGNORECASE
            )
            
            # 보호된 부분들을 다시 복원
            for i, protected in enumerate(protected_parts):
                temp_text = temp_text.replace(f"__PROTECTED_{i}__", protected)
            
            result_text = temp_text
    
    return result_text

async def scroll_and_collect(page):
    """스크롤하면서 실시간 데이터 수집 - 무제한 버전"""
    log_progress("🔄 스크롤 및 데이터 수집 시작... (무제한 모드)")
    
    collected_products = {}  # 중복 제거를 위한 딕셔너리 (key: 상품명+판매처)
    last_height = 0
    no_change_count = 0
    scroll_count = 0
    no_new_products_count = 0
    last_product_count = 0
    
    while True:
        # 현재 보이는 상품들 수집
        await collect_visible_products(page, collected_products)
        
        # 스크롤 전 높이
        current_height = await page.evaluate("document.body.scrollHeight")
        
        # 스크롤 다운
        await page.evaluate("window.scrollBy(0, window.innerHeight * 0.8)")  # 80% 정도만 스크롤
        await page.wait_for_timeout(1500)  # 렌더링 대기
        
        # 스크롤 후 높이
        new_height = await page.evaluate("document.body.scrollHeight")
        
        # 새로운 상품이 추가되었는지 확인
        current_product_count = len(collected_products)
        if current_product_count == last_product_count:
            no_new_products_count += 1
        else:
            no_new_products_count = 0
        last_product_count = current_product_count
        
        # 변화 체크 (페이지 끝 감지를 위해 유지하되, 임계값 증가)
        if new_height == last_height:
            no_change_count += 1
            # 10번 연속 변화가 없고, 새 상품도 없으면 종료
            if no_change_count >= 10 and no_new_products_count >= 10:
                # 마지막까지 스크롤 후 한번 더 수집
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(2000)
                await collect_visible_products(page, collected_products)
                log_progress("✅ 페이지 끝에 도달했습니다.")
                break
        else:
            no_change_count = 0
        
        last_height = new_height
        scroll_count += 1
    
    log_progress(f"✨ 스크롤 완료 ({scroll_count}번) - 총 {len(collected_products)}개 상품 수집")
    return list(collected_products.values())

async def collect_visible_products(page, collected_products):
    """현재 보이는 상품들 수집"""
    product_cards_selector = 'li.compositeCardContainer_composite_card_container__jr8cb.composite_card_container'
    
    try:
        product_cards = await page.query_selector_all(product_cards_selector)
        
        for card in product_cards:
            try:
                product_info = {}
                
                # 상품명 (필수 - 중복 체크용)
                title_element = await card.query_selector('strong.productCardTitle_product_card_title__eQupA')
                if not title_element:
                    continue
                product_info['상품명'] = (await title_element.text_content()).strip()
                
                # 판매처 (중복 체크용)
                mall_element = await card.query_selector('span.productCardMallLink_mall_name__5oWPw')
                product_info['판매처'] = (await mall_element.text_content()).strip() if mall_element else ""
                
                # 유니크 키 생성
                unique_key = f"{product_info['상품명']}_{product_info['판매처']}"
                
                # 이미 수집된 상품이면 스킵
                if unique_key in collected_products:
                    continue
                
                # 썸네일 - 여러 방법으로 시도
                thumbnail_url = await extract_thumbnail(card)
                product_info['썸네일'] = thumbnail_url
                
                # 가격
                price_element = await card.query_selector('span.priceTag_number__1QW0R')
                product_info['가격'] = extract_price_number((await price_element.text_content()).strip()) if price_element else ""
                
                # 배송비
                delivery_badge_element = await card.query_selector('span.productCardDeliveryBadge_text__OrtL_')
                delivery_fee_element = await card.query_selector('span.productCardDeliveryFeeInfo_delivery_text__54pei')
                
                delivery_badge_text = (await delivery_badge_element.text_content()).strip() if delivery_badge_element else ""
                delivery_fee_text = (await delivery_fee_element.text_content()).strip() if delivery_fee_element else ""
                product_info['배송비'] = extract_delivery_fee(delivery_badge_text, delivery_fee_text)
                
                # 수집된 상품 저장
                collected_products[unique_key] = product_info
                
            except Exception as e:
                continue
                
    except Exception as e:
        print(f"⚠️ 상품 수집 중 오류: {e}")

async def extract_thumbnail(card):
    """썸네일 이미지 URL 추출 - 여러 방법 시도"""
    # 방법 1: img 태그의 src 속성
    img_element = await card.query_selector('img.autoFitImg_auto_fit_img__fIpj4, img.productCardThumbnail_image__Li6iz, img[class*="thumbnail"], img[class*="product"]')
    if img_element:
        src = await img_element.get_attribute('src')
        if src and src.startswith('http'):
            return src
    
    # 방법 2: data-src 속성 (lazy loading)
    if img_element:
        data_src = await img_element.get_attribute('data-src')
        if data_src and data_src.startswith('http'):
            return data_src
    
    # 방법 3: srcset 속성
    if img_element:
        srcset = await img_element.get_attribute('srcset')
        if srcset:
            # srcset에서 첫 번째 URL 추출
            urls = re.findall(r'(https?://[^\s,]+)', srcset)
            if urls:
                return urls[0]
    
    # 방법 4: background-image 스타일
    div_with_bg = await card.query_selector('div[style*="background-image"]')
    if div_with_bg:
        style = await div_with_bg.get_attribute('style')
        if style:
            bg_url = re.search(r'url\(["\']?(https?://[^"\']+)["\']?\)', style)
            if bg_url:
                return bg_url.group(1)
    
    # 방법 5: a 태그 내부의 img 찾기
    link_element = await card.query_selector('a.productCardLink_link__bCGy9')
    if link_element:
        img_in_link = await link_element.query_selector('img')
        if img_in_link:
            src = await img_in_link.get_attribute('src')
            if src and src.startswith('http'):
                return src
    
    return ""

def extract_price_number(price_text):
    """가격에서 숫자 추출"""
    if not price_text:
        return ""
    numbers = re.findall(r'[\d,]+', price_text)
    return numbers[0].replace(',', '') if numbers else ""

def extract_delivery_fee(delivery_badge_text, delivery_fee_text):
    """배송비 정보 추출"""
    if delivery_badge_text and delivery_badge_text.strip() in ['무료배송', '멤버십 무료반품 혜택']:
        return "0"
    
    if delivery_fee_text:
        price_match = re.search(r'([\d,]+)원', delivery_fee_text.strip())
        if price_match:
            return price_match.group(1).replace(',', '')
    
    return ""

def log_keyword_matching(products_data, highlight_keywords):
    """키워드 매칭 작업 진행 로그를 상품 단위로 출력하고 매칭된 상품 수를 반환"""
    total = len(products_data)
    matched_products = 0
    if not highlight_keywords:
        print(f"📦 키워드 매칭 완료: 총 {total}개 (키워드 없음)")
        return 0
    lowered_keywords = [kw.strip().lower() for kw in highlight_keywords if kw and kw.strip()]
    for product in products_data:
        try:
            name = str(product.get('상품명', '')).strip().lower()
            mall = str(product.get('판매처', '')).strip().lower()
            for kw in lowered_keywords:
                if kw and (kw in name or kw in mall):
                    matched_products += 1
                    product['매칭'] = True
                    # 원본 보존
                    product['매칭키워드'] = [k for k in highlight_keywords if k and k.strip().lower() == kw]
                    break
            else:
                product['매칭'] = False
                product['매칭키워드'] = []
        except Exception:
            pass
    print(f"🎯 키워드 매칭 완료: 총 {total}개 중 {matched_products}개 매칭")
    return matched_products

async def save_products_data(products_data, highlight_keywords=None, page=None):
    """데이터 저장"""
    if not products_data:
        return None
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # 스크립트와 같은 폴더에 결과물 저장 (CSS는 인라인 처리)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # 결과물 저장 폴더 준비
    results_dir = os.path.join(script_dir, "results")
    try:
        os.makedirs(results_dir, exist_ok=True)
    except Exception:
        pass
    
    # 경로 설정: 모든 산출물을 스크립트 폴더에 저장
    csv_basename = f"naver_shopping_products_{timestamp}.csv"
    csv_filename = os.path.join(results_dir, csv_basename)
    try:
        t0 = perf_counter()
        log_progress("📊 CSV 저장 시작")
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['썸네일', '판매처', '상품명', '가격', '배송비']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            for product in products_data:
                writer.writerow(product)
        log_progress(f"✅ CSV 저장 완료 ({len(products_data)}행, {perf_counter()-t0:.2f}s)")
    except Exception as e:
        print(f"⚠️ CSV 저장 오류: {e}")
    
    # JSON 저장
    json_basename = f"naver_shopping_products_{timestamp}.json"
    json_filename = os.path.join(results_dir, json_basename)
    try:
        t0 = perf_counter()
        log_progress("📄 JSON 저장 시작")
        with open(json_filename, 'w', encoding='utf-8') as jsonfile:
            json.dump(products_data, jsonfile, ensure_ascii=False, indent=2)
        log_progress(f"✅ JSON 저장 완료 ({perf_counter()-t0:.2f}s)")
    except Exception as e:
        print(f"⚠️ JSON 저장 오류: {e}")
    
    # HTML 저장 (키워드 하이라이트 포함)
    html_basename = f"naver_shopping_products_{timestamp}.html"
    html_filename = os.path.join(results_dir, html_basename)
    try:
        t0 = perf_counter()
        log_progress("🎨 HTML 생성 시작")
        # 진행 게이지 표시 (페이지가 있는 경우)
        if page is not None:
            try:
                await create_progress_bar(page, "리포트 구성 시작...")
                await update_progress_bar(page, 2, "리포트 헤더 준비...")
            except Exception:
                pass

        # 무거운 HTML 생성을 별도 스레드에서 수행하며, 스레드-안전 진행 콜백으로 실제 진행률 반영
        progress_cb = None
        if page is not None:
            try:
                loop = asyncio.get_running_loop()
                def progress_cb(pct, text=None):
                    try:
                        loop.call_soon_threadsafe(asyncio.create_task, update_progress_bar(page, int(pct), text))
                    except Exception:
                        pass
            except Exception:
                progress_cb = None

        html_task = asyncio.create_task(
            asyncio.to_thread(
                generate_html_report,
                products_data,
                timestamp,
                highlight_keywords,
                progress_cb,
                None,
            )
        )

        # 콜백이 없는 경우에만 간단한 시뮬레이션으로 사용자 체감 향상
        if page is not None and progress_cb is None:
            pct = 5
            try:
                while not html_task.done():
                    pct = min(99, pct + 1)
                    await update_progress_bar(page, int(pct), f"리포트 구성 중... {int(pct)}%")
                    await asyncio.sleep(0.15)
            except Exception:
                pass

        html_content = await html_task

        if page is not None:
            try:
                await update_progress_bar(page, 98, "리포트 파일 저장 중...")
            except Exception:
                pass
        with open(html_filename, 'w', encoding='utf-8') as htmlfile:
            htmlfile.write(html_content)
        log_progress(f"✅ HTML 저장 완료 ({perf_counter()-t0:.2f}s)")
        # 로그는 results/상대경로 기준으로 안내
        print(f"💎 파일 저장 완료: {os.path.join('results', csv_basename)}, {os.path.join('results', json_basename)}, {os.path.join('results', html_basename)}")
        if highlight_keywords:
            print(f"✨ 키워드 하이라이트 적용됨: {len(highlight_keywords)}개 키워드")
        if page is not None:
            try:
                await update_progress_bar(page, 100, "리포트 저장 완료")
                await finish_progress_bar(page, True)
            except Exception:
                pass
        return html_filename
    except Exception as e:
        print(f"⚠️ HTML 저장 오류: {e}")
        if page is not None:
            try:
                await finish_progress_bar(page, False)
            except Exception:
                pass
        return None

def build_report_css():
    return """
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
        
        :root {
            --premium-dark: #0a0e1a;
            --premium-darker: #050810;
            --premium-surface: #0d1220;
            --premium-card: rgba(16,24,48,0.7);
            --premium-border: rgba(147,197,253,0.08);
            --premium-text: #e8eaed;
            --premium-text-secondary: #94a3b8;
            --premium-gold: #fbbf24;
            --premium-gold-light: #fde68a;
            --premium-accent: #3b82f6;
            --premium-accent-light: #60a5fa;
            --premium-success: #10b981;
            --premium-danger: #ef4444;
            --premium-purple: #8b5cf6;
            --premium-gradient: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            --premium-gradient-gold: linear-gradient(135deg, #f6d365 0%, #fda085 100%);
            --premium-shadow-xl: 0 20px 60px rgba(0,0,0,0.5), 0 0 120px rgba(59,130,246,0.05);
            --premium-shadow-glow: 0 0 40px rgba(59,130,246,0.2);
        }
        
        [data-theme="light"] {
            --premium-dark: #ffffff;
            --premium-darker: #f8fafc;
            --premium-surface: #ffffff;
            --premium-card: rgba(255,255,255,0.95);
            --premium-border: rgba(59,130,246,0.1);
            --premium-text: #0f172a;
            --premium-text-secondary: #64748b;
            --premium-gold: #f59e0b;
            --premium-gold-light: #fbbf24;
            --premium-accent: #2563eb;
            --premium-accent-light: #3b82f6;
            --premium-shadow-xl: 0 20px 60px rgba(0,0,0,0.08), 0 0 120px rgba(59,130,246,0.02);
            --premium-shadow-glow: 0 0 40px rgba(59,130,246,0.1);
        }
        
        * { 
            margin: 0; 
            padding: 0; 
            box-sizing: border-box; 
        }
        
        html, body { 
            height: 100%; 
        }
        
        body {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            background: var(--premium-darker);
            background-image: 
                radial-gradient(ellipse at top left, rgba(59,130,246,0.15), transparent 50%),
                radial-gradient(ellipse at bottom right, rgba(139,92,246,0.15), transparent 50%),
                linear-gradient(180deg, var(--premium-darker) 0%, var(--premium-dark) 100%);
            color: var(--premium-text);
            -webkit-font-smoothing: antialiased;
            -moz-osx-font-smoothing: grayscale;
            padding: 2rem;
            min-height: 100vh;
            position: relative;
            overflow-x: hidden;
        }
        
        body::before {
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: url('data:image/svg+xml,<svg width="100" height="100" xmlns="http://www.w3.org/2000/svg"><defs><pattern id="grid" width="100" height="100" patternUnits="userSpaceOnUse"><path d="M 100 0 L 0 0 0 100" fill="none" stroke="rgba(147,197,253,0.03)" stroke-width="1"/></pattern></defs><rect width="100%" height="100%" fill="url(%23grid)"/></svg>');
            pointer-events: none;
            z-index: 0;
        }
        
        .container { 
            max-width: 1600px; 
            margin: 0 auto; 
            position: relative;
            z-index: 1;
        }
        
        .header {
            background: var(--premium-card);
            backdrop-filter: blur(20px) saturate(180%);
            border: 1px solid var(--premium-border);
            border-radius: 24px;
            padding: 3rem;
            box-shadow: var(--premium-shadow-xl);
            text-align: center;
            position: relative;
            overflow: hidden;
        }
        
        .header::before {
            content: '';
            position: absolute;
            top: -50%;
            left: -50%;
            width: 200%;
            height: 200%;
            background: var(--premium-gradient);
            opacity: 0.03;
            animation: shimmer 20s ease-in-out infinite;
        }
        
        @keyframes shimmer {
            0%, 100% { transform: translateX(-50%) translateY(-50%) rotate(0deg); }
            50% { transform: translateX(-30%) translateY(-30%) rotate(180deg); }
        }
        
        .header h1 { 
            font-size: 2.5rem; 
            font-weight: 800;
            letter-spacing: -0.02em;
            margin-bottom: 0.5rem;
            background: linear-gradient(135deg, var(--premium-text) 0%, var(--premium-accent-light) 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            position: relative;
        }
        
        .header .subtitle { 
            color: var(--premium-text-secondary); 
            font-size: 1rem;
            font-weight: 500;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            opacity: 0.8;
        }
        
        .header .instructions {
            color: var(--premium-text-secondary);
            font-size: 0.875rem;
            margin-top: 1.5rem;
            padding: 1rem;
            background: rgba(59,130,246,0.05);
            border: 1px solid rgba(59,130,246,0.1);
            border-radius: 12px;
            backdrop-filter: blur(10px);
        }
        
        .stats {
            display: flex;
            justify-content: center;
            align-items: center;
            gap: 1rem;
            margin-top: 2rem;
            flex-wrap: wrap;
        }
        
        .stat-item {
            background: linear-gradient(135deg, rgba(59,130,246,0.1), rgba(139,92,246,0.1));
            border: 1px solid rgba(147,197,253,0.2);
            color: var(--premium-text);
            padding: 0.75rem 1.5rem;
            border-radius: 999px;
            font-size: 0.875rem;
            font-weight: 600;
            box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            backdrop-filter: blur(10px);
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        .stat-item:hover {
            transform: translateY(-2px);
            box-shadow: var(--premium-shadow-glow);
        }
        
        .stat-item.highlight {
            background: var(--premium-gradient-gold);
            border-color: var(--premium-gold);
            color: var(--premium-darker);
            font-weight: 700;
        }
        
        .products-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
            gap: 1.5rem;
            margin-top: 3rem;
        }
        
        .product-card {
            background: var(--premium-card);
            backdrop-filter: blur(20px) saturate(180%);
            border: 1px solid var(--premium-border);
            border-radius: 20px;
            overflow: hidden;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            cursor: pointer;
            position: relative;
        }
        
        .product-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: linear-gradient(135deg, transparent, rgba(59,130,246,0.05));
            opacity: 0;
            transition: opacity 0.4s ease;
        }
        
        .product-card:hover {
            transform: translateY(-8px) scale(1.02);
            box-shadow: var(--premium-shadow-glow);
            border-color: var(--premium-accent);
        }
        
        .product-card:hover::before {
            opacity: 1;
        }
        
        .product-card.selected {
            border-color: var(--premium-gold);
            box-shadow: 0 0 0 3px rgba(251,191,36,0.2), var(--premium-shadow-glow);
        }
        
        .product-card.selected::after {
            content: '✓';
            position: absolute;
            top: 1rem;
            right: 1rem;
            background: var(--premium-gradient-gold);
            color: var(--premium-darker);
            width: 2rem;
            height: 2rem;
            border-radius: 999px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-weight: 900;
            z-index: 10;
            animation: checkIn 0.4s cubic-bezier(0.68, -0.55, 0.265, 1.55);
        }
        
        @keyframes checkIn {
            from { transform: scale(0) rotate(-180deg); }
            to { transform: scale(1) rotate(0); }
        }
        
        .product-image {
            width: 100%;
            height: 220px;
            object-fit: contain;
            background: linear-gradient(135deg, rgba(59,130,246,0.03), rgba(139,92,246,0.03));
            padding: 1rem;
        }
        
        .product-info {
            padding: 1.5rem;
        }
        
        .product-mall {
            color: var(--premium-accent-light);
            font-size: 0.75rem;
            margin-bottom: 0.5rem;
            font-weight: 700;
            letter-spacing: 0.05em;
            text-transform: uppercase;
        }
        
        .product-title {
            font-size: 0.95rem;
            font-weight: 600;
            color: var(--premium-text);
            margin-bottom: 1rem;
            line-height: 1.4;
            display: block;
            word-break: break-word;
        }
        
        .product-price {
            font-size: 1.5rem;
            font-weight: 800;
            background: var(--premium-gradient);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 0.5rem;
        }
        
        .product-delivery {
            font-size: 0.75rem;
            color: var(--premium-success);
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }
        
        .product-delivery.paid {
            color: var(--premium-text-secondary);
        }
        
        .no-image {
            width: 100%;
            height: 220px;
            background: linear-gradient(135deg, rgba(59,130,246,0.05), rgba(139,92,246,0.05));
            display: flex;
            align-items: center;
            justify-content: center;
            color: var(--premium-text-secondary);
            font-size: 0.875rem;
        }
        
        .floating-btn {
            position: fixed;
            bottom: 2rem;
            right: 2rem;
            background: var(--premium-gradient);
            color: white;
            padding: 1rem 1.5rem;
            border-radius: 999px;
            font-weight: 700;
            cursor: pointer;
            box-shadow: var(--premium-shadow-glow);
            text-decoration: none;
            z-index: 1000;
            border: 0;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        .floating-btn:hover {
            transform: translateY(-4px) scale(1.05);
            box-shadow: 0 20px 40px rgba(59,130,246,0.4);
        }
        
        .highlight-toggle-btn {
            position: fixed;
            bottom: 5.5rem;
            right: 2rem;
            background: var(--premium-gradient-gold);
            color: var(--premium-darker);
            padding: 1rem 1.5rem;
            border-radius: 999px;
            font-weight: 700;
            cursor: pointer;
            box-shadow: 0 10px 30px rgba(251,191,36,0.3);
            border: none;
            z-index: 1000;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        .highlight-toggle-btn:hover {
            transform: translateY(-4px) scale(1.05);
        }
        
        .highlight-toggle-btn.off {
            opacity: 0.6;
            filter: grayscale(0.5);
        }
        
        .csv-export-btn {
            position: fixed;
            bottom: 9rem;
            right: 2rem;
            background: linear-gradient(135deg, #10b981, #059669);
            color: white;
            padding: 1rem 1.5rem;
            border-radius: 999px;
            font-weight: 700;
            cursor: pointer;
            box-shadow: 0 10px 30px rgba(16,185,129,0.3);
            border: none;
            z-index: 1000;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        .csv-export-btn:hover {
            transform: translateY(-4px) scale(1.05);
        }
        
        .theme-toggle-btn {
            position: fixed;
            bottom: 2rem;
            left: 2rem;
            background: var(--premium-card);
            backdrop-filter: blur(10px);
            border: 1px solid var(--premium-border);
            color: var(--premium-text);
            padding: 1rem;
            border-radius: 999px;
            font-weight: 700;
            cursor: pointer;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            z-index: 1000;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        .theme-toggle-btn:hover {
            transform: translateY(-2px);
            box-shadow: var(--premium-shadow-glow);
        }
        
        .delete-info {
            position: fixed;
            top: 2rem;
            right: 2rem;
            background: var(--premium-card);
            backdrop-filter: blur(20px);
            color: var(--premium-text);
            padding: 1rem 1.5rem;
            border-radius: 16px;
            border: 1px solid var(--premium-border);
            font-weight: 600;
            z-index: 1000;
            display: none;
            box-shadow: var(--premium-shadow-xl);
        }
        
        .filter-controls {
            display: flex;
            gap: 0.5rem;
            justify-content: center;
            margin-top: 1.5rem;
            flex-wrap: wrap;
        }
        
        .filter-btn {
            background: rgba(59,130,246,0.1);
            color: var(--premium-text);
            border: 1px solid rgba(59,130,246,0.2);
            padding: 0.5rem 1rem;
            border-radius: 999px;
            cursor: pointer;
            font-weight: 600;
            font-size: 0.75rem;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        
        .filter-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 20px rgba(59,130,246,0.3);
        }
        
        .filter-btn.selected {
            background: var(--premium-gradient);
            border-color: transparent;
            color: white;
            box-shadow: var(--premium-shadow-glow);
        }
        
        .highlight {
            background: linear-gradient(transparent 60%, rgba(251,191,36,0.4) 60%);
            color: inherit !important;
            font-weight: 700 !important;
            padding: 0 2px;
            border-radius: 2px;
        }
        
        .no-highlight .highlight {
            background: none !important;
            font-weight: inherit !important;
        }
        
        .product-card.matched {
            border-color: var(--premium-gold);
            box-shadow: 0 0 0 2px rgba(251,191,36,0.2);
        }
        
        @keyframes fadeOut {
            from {
                opacity: 1;
                transform: scale(1) rotate(0deg);
            }
            to {
                opacity: 0;
                transform: scale(0.8) rotate(10deg);
            }
        }
        
        @media (max-width: 860px) {
            .csv-export-btn { right: 1.5rem; bottom: 9rem; }
            .highlight-toggle-btn { right: 1.5rem; bottom: 5.5rem; }
            .floating-btn { right: 1.5rem; bottom: 2rem; }
            .products-grid { grid-template-columns: repeat(auto-fill, minmax(240px, 1fr)); }
        }
    """

def generate_html_report(products_data, timestamp, highlight_keywords=None, progress_callback=None, css_filename=None):
    """HTML 리포트 생성 - 모든 상품 표시 (선택 삭제 기능 및 키워드 하이라이트 포함)"""
    total_products = len(products_data)
    current_time = datetime.now().strftime("%Y년 %m월 %d일 %H:%M:%S")
    # 진행 로그 주기 제거(요약만 유지)
    if progress_callback:
        try:
            progress_callback(5, "리포트 헤더 구성 중...")
        except Exception:
            pass
    
    # 이미지 있는 상품 수 계산
    products_with_image = sum(1 for p in products_data if p.get('썸네일'))
    
    # 키워드 매칭된 상품 수 계산
    matched_products = 0
    if highlight_keywords:
        for product in products_data:
            product_name = product.get('상품명', '')
            mall_name = product.get('판매처', '')
            for keyword in highlight_keywords:
                if keyword.lower() in product_name.lower() or keyword.lower() in mall_name.lower():
                    matched_products += 1
                    break
    
    # 필터 컨트롤 HTML (백슬래시 이스케이프가 필요한 onclick을 f-string 밖에서 구성)
    filter_controls = ""
    if highlight_keywords:
        filter_controls = (
            '<div class="filter-controls">\n'
            '    <button class="filter-btn selected" id="filter-all" onclick="setFilterMode(\'all\')">전체 상품</button>\n'
            '    <button class="filter-btn" id="filter-on" onclick="setFilterMode(\'on\')">매칭 상품</button>\n'
            '    <button class="filter-btn" id="filter-off" onclick="setFilterMode(\'off\')">미매칭 상품</button>\n'
            '</div>'
        )

    # 추가 콘솔 안내/초기 필터 적용 JS (f-string 내부에서 직접 조건식을 쓰지 않도록 사전 구성)
    extra_console_tip = "console.log('⌨️ 1/2/3 키로 전체/매칭/미매칭 필터링');" if highlight_keywords else ""
    extra_init_filter = "setFilterMode('all');" if highlight_keywords else ""

    # CSS 인라인 또는 링크 결정 (기본: 인라인)
    css_block = (
        f"<style>\n{build_report_css()}\n</style>" if not css_filename else f'<link rel="stylesheet" href="{css_filename}">'
    )

    html = f"""<!DOCTYPE html>
<html lang=\"ko\" data-theme=\"dark\">
<head>
    <meta charset=\"UTF-8\">
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
    <title>Premium Shopping Analytics - {current_time}</title>
    {css_block}
</head>
<body>
    <div class=\"container\">
        <div class=\"header\">
            <h1>💎 Premium Shopping Analytics</h1>
            <div class=\"subtitle\">Naver Shopping Intelligence Report</div>
            <div class=\"instructions\">
                ✨ 프리미엄 데이터 분석 도구 | 클릭으로 선택 → Delete로 제거 | 키워드 하이라이트 지원
                <br>⌨️ H: 하이라이트 토글 · ESC: 선택 해제 · Ctrl+A: 전체 선택
            </div>
            <div class=\"stats\"> 
                <div class=\"stat-item\" id=\"total-count\">총 {total_products}개 상품</div>
                <div class=\"stat-item\">이미지 수집 {products_with_image}개</div>"""
    
    if highlight_keywords:
        html += f"""
                <div class="stat-item highlight">매칭 {matched_products}개</div>
                <div class="stat-item highlight">키워드 {len(highlight_keywords)}개</div>"""
    
    html += f"""
                <div class=\"stat-item\">Report ID: {timestamp}</div>
            </div>
            {filter_controls}
        </div>
        <div class=\"delete-info\" id=\"delete-info\">선택: <span id=\"selected-count\">0</span>개 | Delete로 삭제</div>
        <div class=\"products-grid\" id=\"products-grid\">
"""
    
    # 상품 카드 구성 진행률 계산용 (최대 약 50회 업데이트)
    step = max(1, total_products // 50) if total_products else 1
    for i, product in enumerate(products_data):
        price = product.get('가격', '')
        formatted_price = f"{int(price):,}원" if price else "가격 정보 없음"
        
        delivery = product.get('배송비', '')
        if delivery == "0" or delivery == "":
            delivery_text = "무료배송"
            delivery_class = "product-delivery"
        else:
            delivery_text = f"배송비 {int(delivery):,}원"
            delivery_class = "product-delivery paid"
        
        # 키워드 하이라이트 적용
        product_name = product.get('상품명', '')
        mall_name = product.get('판매처', '')
        
        is_matched = False
        if highlight_keywords:
            highlighted_name = highlight_keywords_in_text(product_name, highlight_keywords)
            highlighted_mall = highlight_keywords_in_text(mall_name, highlight_keywords)
            
            # 매칭 여부 확인
            for keyword in highlight_keywords:
                if keyword.lower() in product_name.lower() or keyword.lower() in mall_name.lower():
                    is_matched = True
                    break
        else:
            highlighted_name = product_name
            highlighted_mall = mall_name
        
        # 매칭된 상품에 클래스 추가
        card_class = "product-card matched" if is_matched else "product-card"
        
        thumbnail = product.get('썸네일', '')
        if thumbnail:
            img_element = f'<img src="{thumbnail}" alt="{product_name}" class="product-image" onerror="this.style.display=\'none\'; this.nextElementSibling.style.display=\'flex\';">'
            no_img_element = '<div class="no-image" style="display:none;">이미지 없음</div>'
        else:
            img_element = ''
            no_img_element = '<div class="no-image">이미지 없음</div>'
        
        html += f"""
            <div class="{card_class}" data-product-id="{i}" data-matched="{1 if is_matched else 0}" data-price="{price or ''}" data-delivery="{delivery}" onclick="toggleSelection(this)">
                {img_element}
                {no_img_element}
                <div class="product-info">
                    <div class="product-mall">{highlighted_mall}</div>
                    <div class="product-title">{highlighted_name}</div>
                    <div class="product-price">{formatted_price}</div>
                    <div class="{delivery_class}">{delivery_text}</div>
                </div>
            </div>
        """
        if progress_callback and ((i + 1) % step == 0 or (i + 1) == total_products):
            try:
                # 5%에서 시작하여 최대 99%까지 실제 진행률 반영
                pct = 5 + int(((i + 1) / max(1, total_products)) * 94)
                progress_callback(pct, f"상품 카드 구성 {i + 1}/{total_products}")
            except Exception:
                pass
    # 진행중 상세 로그는 제거됨
    
    html += """
        </div>
    </div>
    <a href=\"#\" class=\"floating-btn\" onclick=\"window.scrollTo({top: 0, behavior: 'smooth'}); return false;\">↑ TOP</a>"""
    
    # 키워드가 있는 경우에만 하이라이트 토글 버튼 추가
    if highlight_keywords:
        html += f"""
    <button class=\"highlight-toggle-btn\" id=\"highlight-toggle\" onclick=\"toggleHighlight()\">✨ 하이라이트</button>"""
    
    html += f"""
    <button class=\"csv-export-btn\" id=\"csv-export\" onclick=\"exportVisibleToCSV()\">📊 CSV 내보내기</button>
    <button class=\"theme-toggle-btn\" id=\"theme-toggle\" onclick=\"toggleTheme()\">🌙</button>
    
    <script>
        let selectedCards = new Set();
        let highlightEnabled = true;
    let filterMode = 'all'; // 'all' | 'on' | 'off'
        // 테마 토글 유지
        (function(){{
            try {{
                const saved = localStorage.getItem('theme');
                if (saved === 'dark' || saved === 'light') {{
                    document.documentElement.setAttribute('data-theme', saved);
                }}
            }} catch (e) {{}}
        }})();

        // 화면 가시성 판단 함수들
        function isCardVisible(card) {{
            // display:none 이거나 DOM에서 보이지 않으면 제외
            return card && card.offsetParent !== null && card.style.display !== 'none';
        }}

        // 현재 페이지(필터 적용 후 보이는 모든 카드) 전체 선택
        function selectAllPageCards() {{
            const cards = document.querySelectorAll('.product-card');
            let changed = 0;
            cards.forEach(card => {{
                // 필터로 숨겨진 카드(display:none)는 제외하고, 실제로 보이는 카드만 선택
                if (isCardVisible(card) && card.style.display !== 'none') {{
                    const productId = card.getAttribute('data-product-id');
                    if (!selectedCards.has(productId)) {{
                        selectedCards.add(productId);
                        card.classList.add('selected');
                        changed++;
                    }}
                }}
            }});
            if (changed > 0) {{
                updateSelectionInfo();
            }}
        }}
        
        function toggleHighlight() {{
            highlightEnabled = !highlightEnabled;
            const toggleBtn = document.getElementById('highlight-toggle');
            const body = document.body;

            if (highlightEnabled) {{
                // 하이라이트 켜기: 텍스트는 그대로 두고 스타일만 복원
                body.classList.remove('no-highlight');
                // 매칭 카드 테두리 복원
                document.querySelectorAll('.product-card[data-matched="1"]').forEach(card => card.classList.add('matched'));
                toggleBtn.textContent = '✨ 하이라이트';
                toggleBtn.classList.remove('off');
            }} else {{
                // 하이라이트 끄기: 텍스트 유지, 스타일만 제거
                body.classList.add('no-highlight');
                // 매칭 카드 테두리 제거
                document.querySelectorAll('.product-card[data-matched="1"]').forEach(card => card.classList.remove('matched'));
                toggleBtn.textContent = '◯ 하이라이트';
                toggleBtn.classList.add('off');
            }}
        }}
        
        function toggleSelection(card) {{
            const productId = card.getAttribute('data-product-id');
            
            if (selectedCards.has(productId)) {{
                selectedCards.delete(productId);
                card.classList.remove('selected');
            }} else {{
                selectedCards.add(productId);
                card.classList.add('selected');
            }}
            
            updateSelectionInfo();
        }}
        
        function updateSelectionInfo() {{
            const selectedCount = selectedCards.size;
            const deleteInfo = document.getElementById('delete-info');
            const selectedCountSpan = document.getElementById('selected-count');
            
            if (selectedCount > 0) {{
                deleteInfo.style.display = 'block';
                selectedCountSpan.textContent = selectedCount;
            }} else {{
                deleteInfo.style.display = 'none';
            }}
        }}
        
        function deleteSelectedCards() {{
            if (selectedCards.size === 0) {{
                return;
            }}
            
            // 선택된 카드들을 삭제 (확인 과정 없이 바로 삭제)
            selectedCards.forEach(productId => {{
                const card = document.querySelector(`[data-product-id="${{productId}}"]`);
                if (card) {{
                    card.style.animation = 'fadeOut 0.3s ease-out';
                    setTimeout(() => {{
                        card.remove();
                        updateTotalCount();
                    }}, 300);
                }}
            }});
            
            selectedCards.clear();
            updateSelectionInfo();
        }}
        
        function updateTotalCount() {{
            // 보이는 카드만 집계
            const remainingCards = Array.from(document.querySelectorAll('.product-card')).filter(c => c.style.display !== 'none').length;
            const totalCountElement = document.getElementById('total-count');
            totalCountElement.textContent = `총 ${{remainingCards}}개 상품`;
        }}

        function setFilterMode(mode) {{
            filterMode = mode;
            // 버튼 선택 상태 업데이트
            const btnAll = document.getElementById('filter-all');
            const btnOn = document.getElementById('filter-on');
            const btnOff = document.getElementById('filter-off');
            [btnAll, btnOn, btnOff].forEach(btn => btn && btn.classList.remove('selected'));
            if (mode === 'all' && btnAll) btnAll.classList.add('selected');
            if (mode === 'on' && btnOn) btnOn.classList.add('selected');
            if (mode === 'off' && btnOff) btnOff.classList.add('selected');
            applyFilter();
        }}

        function applyFilter() {{
            const cards = document.querySelectorAll('.product-card');
            cards.forEach(card => {{
                const matched = card.getAttribute('data-matched') === '1';
                if (filterMode === 'all') {{
                    card.style.display = '';
                }} else if (filterMode === 'on') {{
                    card.style.display = matched ? '' : 'none';
                }} else if (filterMode === 'off') {{
                    card.style.display = matched ? 'none' : '';
                }}
            }});
            updateTotalCount();
        }}
        
        // 키보드 단축키 이벤트 리스너
        document.addEventListener('keydown', function(event) {{
            // 입력 필드/편집 영역에서는 기본 Ctrl+A 동작을 유지
            const tag = (event.target && event.target.tagName || '').toLowerCase();
            const isEditable = event.target && (event.target.isContentEditable || tag === 'input' || tag === 'textarea' || tag === 'select');
            
            // Ctrl+A: 현재 페이지의 모든 상품 전체 선택(필터 적용)
            if (!isEditable && event.ctrlKey && (event.key === 'a' || event.key === 'A')) {{
                event.preventDefault();
                event.stopPropagation();
                selectAllPageCards();
                return;
            }}
            if (event.key === 'Delete' || event.key === 'Del') {{
                deleteSelectedCards();
            }}
            
            // Escape 키로 선택 해제
            if (event.key === 'Escape') {{
                selectedCards.forEach(productId => {{
                    const card = document.querySelector(`[data-product-id="${{productId}}"]`);
                    if (card) {{
                        card.classList.remove('selected');
                    }}
                }});
                selectedCards.clear();
                updateSelectionInfo();
            }}
            
            // H 키로 하이라이트 토글 (키워드가 있는 경우에만)
            if (event.key === 'h' || event.key === 'H') {{
                const toggleBtn = document.getElementById('highlight-toggle');
                if (toggleBtn) {{
                    toggleHighlight();
                }}
            }}

            // 1/2/3 단축키로 필터 변경 (키워드 있을 때만 동작)
            if ({'true' if highlight_keywords else 'false'}) {{
                if (event.key === '1') setFilterMode('all');
                if (event.key === '2') setFilterMode('on');
                if (event.key === '3') setFilterMode('off');
            }}
        }});
        
        function toggleTheme() {{
            try {{
                const el = document.documentElement;
                const cur = el.getAttribute('data-theme') || 'dark';
                const next = cur === 'light' ? 'dark' : 'light';
                el.setAttribute('data-theme', next);
                document.getElementById('theme-toggle').textContent = next === 'light' ? '☀️' : '🌙';
                try {{ localStorage.setItem('theme', next); }} catch (e) {{}}
            }} catch (e) {{}}
        }}
        
        // 초기화 시 사용법 안내
        console.log('💎 Premium Shopping Analytics');
        console.log('━━━━━━━━━━━━━━━━━━━━━━━━━━━');
        console.log('📍 상품 카드를 클릭하여 선택/해제');
        console.log('🗑️ Delete 키로 선택된 상품 삭제');
        console.log('⎋ Escape 키로 모든 선택 해제');
        console.log('✨ H 키로 하이라이트 온/오프');
        console.log('⌨️ Ctrl+A로 현재 페이지 전체 선택');
        {extra_console_tip}

        // 초기 필터 적용 (키워드가 있는 경우에만 컨트롤 보임)
        {extra_init_filter}

        // 초기 테마 아이콘 설정
        (function() {{
            const theme = document.documentElement.getAttribute('data-theme') || 'dark';
            document.getElementById('theme-toggle').textContent = theme === 'light' ? '☀️' : '🌙';
        }})();

        // CSV 다운로드: 현재 보이는(필터/삭제 반영) 상품들만 내보내기
        function exportVisibleToCSV() {{
            try {{
                const cards = Array.from(document.querySelectorAll('.product-card'))
                    .filter(function(c) {{ return c.style.display !== 'none'; }});
                if (cards.length === 0) {{
                    alert('내보낼 상품이 없습니다.');
                    return;
                }}
                // 헤더
                const rows = [['상품명', '전체가격(판매가+배송비)']];
                const getText = function(el, sel) {{
                    const n = el.querySelector(sel);
                    return n ? (n.textContent || '').trim() : '';
                }};
                cards.forEach(function(card) {{
                    const name = getText(card, '.product-title');
                    const priceNum = parseInt(card.getAttribute('data-price') || '0', 10) || 0;
                    const deliveryNum = parseInt(card.getAttribute('data-delivery') || '0', 10) || 0;
                    const total = priceNum + deliveryNum;
                    rows.push([name, String(total)]);
                }});
                // CSV 인코딩: 정규식 없이 안전하게(쉼표/따옴표/개행 포함 시 따옴표로 감싸고 내부 따옴표 이스케이프)
                const esc = function(field) {{
                    const s = String((field === null || field === undefined) ? '' : field);
                    const needsQuote = (s.indexOf('"') !== -1) || (s.indexOf(',') !== -1) || (s.indexOf('\\n') !== -1) || (s.indexOf('\\r') !== -1);
                    const doubled = s.split('"').join('""');
                    return needsQuote ? '"' + doubled + '"' : doubled;
                }};
                const csv = rows.map(function(r) {{ return r.map(esc).join(','); }}).join('\\r\\n');
                // UTF-8 BOM 추가하여 Excel에서 한글/CSV 인코딩 문제 해결
                const bom = new Uint8Array([0xEF, 0xBB, 0xBF]);
                const blob = new Blob([bom, csv], {{ type: 'text/csv;charset=utf-8;' }});
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                var now = new Date();
                var pad = function(n) {{ return (n < 10 ? '0' + n : n); }};
                var ts = '' + now.getFullYear() + pad(now.getMonth() + 1) + pad(now.getDate()) + '_' + pad(now.getHours()) + pad(now.getMinutes()) + pad(now.getSeconds());
                a.href = url;
                a.download = 'premium_products_' + ts + '.csv';
                document.body.appendChild(a);
                a.click();
                setTimeout(function() {{
                    URL.revokeObjectURL(url);
                    a.remove();
                }}, 0);
            }} catch (e) {{
                console.error('CSV 내보내기 실패:', e);
                alert('CSV 내보내기 중 오류가 발생했습니다. 콘솔을 확인하세요.');
            }}
        }}
    </script>
</body>
</html>
"""
    return html

async def create_ready_button(page):
    """준비완료 버튼 생성 - 프리미엄 디자인"""
    return await page.evaluate("""
        (() => {
            try {
                const existingBtns = document.querySelectorAll('#crawling-ready-btn, #crawling-notice');
                existingBtns.forEach(el => el.remove());
                
                const button = document.createElement('button');
                button.innerHTML = '<span style="display: inline-block; animation: pulse 2s infinite;">💎</span> Premium Crawler Ready';
                button.id = 'crawling-ready-btn';
                button.style.cssText = `
                    position: fixed !important; 
                    top: 30px !important; 
                    left: 50% !important;
                    transform: translateX(-50%) !important; 
                    z-index: 2147483647 !important;
                    padding: 18px 48px !important; 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
                    color: white !important; 
                    border: 2px solid rgba(255,255,255,0.2) !important; 
                    border-radius: 50px !important;
                    font-size: 16px !important; 
                    font-weight: 700 !important; 
                    cursor: pointer !important;
                    box-shadow: 0 10px 40px rgba(102,126,234,0.4), 0 0 0 3px rgba(102,126,234,0.1) !important;
                    display: block !important; 
                    min-width: 320px !important; 
                    text-align: center !important;
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
                    letter-spacing: 0.5px !important;
                    text-transform: uppercase !important;
                    backdrop-filter: blur(10px) !important;
                    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
                `;
                
                // 애니메이션 스타일 추가
                const style = document.createElement('style');
                style.textContent = `
                    @keyframes pulse {
                        0% { transform: scale(1); }
                        50% { transform: scale(1.2); }
                        100% { transform: scale(1); }
                    }
                    #crawling-ready-btn:hover {
                        transform: translateX(-50%) translateY(-3px) !important;
                        box-shadow: 0 15px 50px rgba(102,126,234,0.5), 0 0 0 5px rgba(102,126,234,0.15) !important;
                    }
                `;
                document.head.appendChild(style);
                
                document.body.insertBefore(button, document.body.firstChild);
                
                button.onclick = function(event) {
                    this.innerHTML = '<span style="animation: spin 1s linear infinite;">⚡</span> Initiating Premium Scan...';
                    this.disabled = true;
                    this.style.background = 'linear-gradient(135deg, #f093fb 0%, #f5576c 100%) !important';
                    this.style.cursor = 'not-allowed !important';
                    this.style.opacity = '0.9 !important';
                    
                    // 스핀 애니메이션 추가
                    const spinStyle = document.createElement('style');
                    spinStyle.textContent = `
                        @keyframes spin {
                            from { transform: rotate(0deg); }
                            to { transform: rotate(360deg); }
                        }
                    `;
                    document.head.appendChild(spinStyle);
                    
                    window.crawlingReady = true;
                    event.preventDefault();
                };
                
                return true;
            } catch (error) {
                return false;
            }
        })()
    """)

async def monitor_and_recreate_button(page):
    """페이지 모니터링 및 버튼 재생성"""
    current_url = page.url
    button_ready = False
    
    while not button_ready:
        try:
            new_url = page.url
            if new_url != current_url:
                current_url = new_url
                await page.wait_for_load_state('domcontentloaded')
                await page.wait_for_timeout(2000)
                await create_ready_button(page)
            
            button_ready = await page.evaluate("window.crawlingReady === true")
            
            if not button_ready:
                button_exists = await page.evaluate("!!document.getElementById('crawling-ready-btn')")
                if not button_exists:
                    await create_ready_button(page)
            
            await page.wait_for_timeout(1000)
            
        except Exception:
            await page.wait_for_timeout(1000)
    
    return True

async def wait_for_user_ready(page):
    """사용자 준비 대기"""
    await page.wait_for_load_state('domcontentloaded')
    await page.wait_for_timeout(2000)
    
    await create_ready_button(page)
    print("💎 화면 상단의 Premium Crawler 버튼을 클릭하여 크롤링을 시작하세요!")
    
    await monitor_and_recreate_button(page)
    
    await page.wait_for_timeout(1000)
    await page.evaluate("""
        (() => {
            const button = document.getElementById('crawling-ready-btn');
            if (button) button.remove();
        })()
    """)

async def create_progress_bar(page, initial_text="리포트 구성 준비 중..."):
    """상단 중앙(주황 버튼 자리)에 진행 게이지 생성 - 프리미엄 디자인"""
    try:
        return await page.evaluate(
            """
            (text) => {
                try {
                    // 기존 버튼/게이지 정리
                    const btn = document.getElementById('crawling-ready-btn');
                    if (btn) btn.remove();
                    const existing = document.getElementById('crawling-progress');
                    if (existing) return true;

                    const container = document.createElement('div');
                    container.id = 'crawling-progress';
                    container.style.cssText = `
                        position: fixed; 
                        top: 30px; 
                        left: 50%; 
                        transform: translateX(-50%);
                        z-index: 2147483647; 
                        width: 480px; 
                        padding: 20px 24px;
                        background: rgba(10,14,26,0.95); 
                        border-radius: 20px;
                        border: 1px solid rgba(147,197,253,0.2); 
                        box-shadow: 0 20px 60px rgba(102,126,234,0.3), 0 0 0 1px rgba(102,126,234,0.1);
                        backdrop-filter: blur(20px) saturate(180%);
                    `;

                    const label = document.createElement('div');
                    label.id = 'crawling-progress-label';
                    label.textContent = text || '리포트 구성 진행 중...';
                    label.style.cssText = `
                        font-weight: 600; 
                        color: #e8eaed; 
                        margin-bottom: 12px; 
                        text-align: center;
                        font-size: 14px;
                        letter-spacing: 0.5px;
                        text-transform: uppercase;
                    `;

                    const track = document.createElement('div');
                    track.style.cssText = `
                        width: 100%; 
                        height: 8px; 
                        background: rgba(147,197,253,0.1); 
                        border-radius: 10px; 
                        overflow: hidden;
                        position: relative;
                    `;

                    const bar = document.createElement('div');
                    bar.id = 'crawling-progress-bar';
                    bar.style.cssText = `
                        width: 0%; 
                        height: 100%; 
                        background: linear-gradient(90deg, #667eea, #764ba2, #f093fb);
                        background-size: 200% 100%;
                        animation: gradientShift 3s ease infinite;
                        transition: width 300ms cubic-bezier(0.4, 0, 0.2, 1);
                        border-radius: 10px;
                        box-shadow: 0 0 20px rgba(102,126,234,0.5);
                    `;

                    // 애니메이션 추가
                    const style = document.createElement('style');
                    style.textContent = `
                        @keyframes gradientShift {
                            0% { background-position: 0% 50%; }
                            50% { background-position: 100% 50%; }
                            100% { background-position: 0% 50%; }
                        }
                    `;
                    document.head.appendChild(style);

                    const percentage = document.createElement('div');
                    percentage.id = 'crawling-progress-percentage';
                    percentage.style.cssText = `
                        text-align: center;
                        color: #94a3b8;
                        font-size: 12px;
                        margin-top: 8px;
                        font-weight: 500;
                    `;
                    percentage.textContent = '0%';

                    track.appendChild(bar);
                    container.appendChild(label);
                    container.appendChild(track);
                    container.appendChild(percentage);
                    document.body.appendChild(container);
                    return true;
                } catch (e) {
                    return false;
                }
            }
            """,
            initial_text,
        )
    except Exception:
        return False

async def update_progress_bar(page, percent, text=None):
    """진행 게이지 업데이트 (퍼센트/라벨) - 프리미엄 디자인"""
    try:
        pct = int(max(0, min(100, int(percent))))
        data = {"pct": pct, "text": text}
        await page.evaluate(
            """
            (data) => {
                const el = document.getElementById('crawling-progress-bar');
                const label = document.getElementById('crawling-progress-label');
                const percentage = document.getElementById('crawling-progress-percentage');
                
                if (el && typeof data.pct === 'number') {
                    el.style.width = `${data.pct}%`;
                }
                if (label && data.text) {
                    label.textContent = data.text;
                }
                if (percentage) {
                    percentage.textContent = `${data.pct}%`;
                }
            }
            """,
            data,
        )
    except Exception:
        pass

async def finish_progress_bar(page, success=True):
    """진행 게이지 완료 표시 후 제거 - 프리미엄 디자인"""
    try:
        await page.evaluate(
            """
            (ok) => {
                const el = document.getElementById('crawling-progress');
                const bar = document.getElementById('crawling-progress-bar');
                const label = document.getElementById('crawling-progress-label');
                const percentage = document.getElementById('crawling-progress-percentage');
                
                if (!el) return;
                
                if (bar) {
                    bar.style.width = '100%';
                    if (ok) {
                        bar.style.background = 'linear-gradient(90deg, #10b981, #059669)';
                        bar.style.animation = 'none';
                    } else {
                        bar.style.background = 'linear-gradient(90deg, #ef4444, #dc2626)';
                        bar.style.animation = 'none';
                    }
                }
                if (label) {
                    label.textContent = ok ? '✨ 리포트 구성 완료' : '⚠️ 리포트 구성 실패';
                    label.style.color = ok ? '#10b981' : '#ef4444';
                }
                if (percentage) {
                    percentage.textContent = '100%';
                }
                
                setTimeout(() => { 
                    el.style.transition = 'opacity 0.5s ease';
                    el.style.opacity = '0';
                    setTimeout(() => {
                        try { el.remove(); } catch(e) {}
                    }, 500);
                }, 1500);
            }
            """,
            True if success else False,
        )
    except Exception:
        pass

def get_chrome_user_data_path():
    """Chrome 사용자 데이터 경로"""
    if sys.platform == "win32":
        user_data_dir = os.path.expanduser("~/AppData/Local/Google/Chrome/User Data")
    elif sys.platform == "darwin":
        user_data_dir = os.path.expanduser("~/Library/Application Support/Google/Chrome")
    else:
        user_data_dir = os.path.expanduser("~/.config/google-chrome")
    
    if not os.path.exists(user_data_dir):
        print(f"⚠️ Chrome 사용자 데이터 디렉토리를 찾을 수 없습니다: {user_data_dir}")
        return None
    
    return user_data_dir

# ===== CDP 자동 실행 유틸 =====
def find_chrome_executable():
    # 우선순위: 환경변수 → 일반 설치 경로 → PATH 검색
    cand = [
        os.getenv('CHROME_EXE'),
        r"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
        r"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe",
    ]
    for c in cand:
        if c and os.path.isfile(c):
            return c
    # PATH
    for name in ("chrome.exe", "chrome"):
        p = shutil.which(name)
        if p:
            return p
    return None

def find_free_port(start=9222, tries=50):
    port = start
    for _ in range(tries):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(("127.0.0.1", port))
                return port
            except OSError:
                port += 1
    return None

def launch_chrome_with_cdp(chrome_path, user_data_dir, profile_dir_name, port):
    args = [
        chrome_path,
        f"--remote-debugging-port={port}",
        "--no-first-run",
        "--no-default-browser-check",
    ]
    if user_data_dir:
        args.append(f"--user-data-dir={user_data_dir}")
    if profile_dir_name:
        args.append(f"--profile-directory={profile_dir_name}")
    # 사용자 체감 위해 크기 고정(최대화는 OS에 따라 무시될 수 있음)
    args.extend(["--start-maximized", "--window-size=1920,1080", "about:blank"])
    # 표준 출력은 필요 없어 숨김
    try:
        proc = subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return proc
    except Exception:
        return None

def wait_cdp_ready(port, timeout=8.0):
    # /json/version에 접근해 wsEndpoint 확인 (네트워크 미사용 환경이면 단순 대기)
    import time as _t
    import urllib.request as _rq
    import urllib.error as _er
    deadline = _t.time() + timeout
    url = f"http://127.0.0.1:{port}/json/version"
    last_err = None
    while _t.time() < deadline:
        try:
            with _rq.urlopen(url, timeout=1.0) as resp:
                if resp.status == 200:
                    return True
        except (_er.URLError, _er.HTTPError, TimeoutError) as e:
            last_err = e
        except Exception as e:
            last_err = e
        _t.sleep(0.25)
    return False

def open_html_result(html_filename):
    """HTML 파일 브라우저로 열기"""
    if html_filename and os.path.exists(html_filename):
        try:
            abs_path = os.path.abspath(html_filename)
            file_url = f"file:///{abs_path.replace(os.sep, '/')}"
            webbrowser.open(file_url)
            print("💎 브라우저에서 프리미엄 리포트를 확인하세요!")
        except Exception as e:
            print(f"⚠️ 브라우저 열기 실패. 수동으로 열어주세요: {html_filename}")

async def access_naver_shopping_optimized():
    """네이버 쇼핑 크롤링 실행"""
    print("\n" + "=" * 70)
    print("💎 Premium Shopping Intelligence System")
    print("=" * 70)
    print("📊 Step 1: 엑셀 파일 키워드 추출")
    print("=" * 70 + "\n")
    
    # 엑셀 파일 선택 및 키워드 추출
    excel_file_path = select_excel_file()
    highlight_keywords = []
    
    if excel_file_path:
        print(f"✅ 선택된 파일: {excel_file_path}")
        highlight_keywords = extract_red_background_cells(excel_file_path)
        
        if highlight_keywords:
            print(f"\n🎯 {len(highlight_keywords)}개의 키워드를 추출했습니다!")
        else:
            print("⚠️ 빨간색 배경 셀을 찾을 수 없습니다.")
            proceed = input("키워드 없이 계속 진행하시겠습니까? (y/n): ")
            if proceed.lower() != 'y':
                print("작업을 취소합니다.")
                return
    else:
        print("⚠️ 엑셀 파일을 선택하지 않았습니다.")
        proceed = input("키워드 없이 계속 진행하시겠습니까? (y/n): ")
        if proceed.lower() != 'y':
            print("작업을 취소합니다.")
            return
    
    print("\n" + "=" * 70)
    print("🚀 Step 2: Premium Crawler 시작")
    print("=" * 70 + "\n")
    
    # 프로필 전략: local(기본) | system | smart(system 실패 시 local)
    def get_or_create_local_chrome_profile():
        try:
            override = os.getenv('CHROME_USER_DATA_DIR')
            if override:
                os.makedirs(override, exist_ok=True)
                return override
        except Exception:
            pass
        script_dir = os.path.dirname(os.path.abspath(__file__))
        profile_dir = os.path.join(script_dir, 'chrome-user-data')
        try:
            os.makedirs(profile_dir, exist_ok=True)
        except Exception:
            pass
        return profile_dir

    # 전략/프로필 디렉터리 결정
    profile_strategy = (os.getenv('PROFILE_STRATEGY', 'smart') or 'smart').strip().lower()
    chrome_profile_dir_name = (os.getenv('CHROME_PROFILE_DIR', 'Default') or 'Default').strip()
    system_user_data_root = get_chrome_user_data_path() if profile_strategy in ('system', 'smart') else None
    if profile_strategy == 'system' and not system_user_data_root:
        print("⚠️ 시스템 Chrome 프로필 경로를 찾지 못해 local 전략으로 전환합니다.")
        profile_strategy = 'local'

    if profile_strategy == 'system':
        user_data_dir = system_user_data_root
        print(f"🔧 프로필 전략: system | User Data: {user_data_dir} | Profile: {chrome_profile_dir_name}")
    else:
        user_data_dir = get_or_create_local_chrome_profile()
        print(f"🔧 프로필 전략: {profile_strategy or 'local'} | User Data: {user_data_dir}")
    
    log_progress("🌐 Playwright 컨텍스트 시작")
    browser = None
    auto_cdp_proc = None
    
    async with async_playwright() as p:
        try:
            print("🔷 Chrome 실행 중...")
            
            # 자동 CDP 시도
            try:
                chrome_exe = find_chrome_executable()
                if chrome_exe:
                    if profile_strategy == 'system':
                        cdp_user_data = system_user_data_root
                        cdp_profile_dir = chrome_profile_dir_name
                    else:
                        cdp_user_data = user_data_dir
                        cdp_profile_dir = None if profile_strategy == 'local' else chrome_profile_dir_name
                    port = find_free_port(9222, 50) or 9222
                    auto_cdp_proc = launch_chrome_with_cdp(chrome_exe, cdp_user_data, cdp_profile_dir, port)
                    if auto_cdp_proc and wait_cdp_ready(port, timeout=8.0):
                        browser = await p.chromium.connect_over_cdp(f"http://127.0.0.1:{port}")
                        context_list = browser.contexts
                        if context_list:
                            page = await context_list[0].new_page()
                        else:
                            context = await browser.new_context()
                            page = await context.new_page()
                        await page.set_viewport_size({'width': 1920, 'height': 1080})
                        print(f"✅ 자동 CDP 연결 성공: 포트 {port}")
                    else:
                        if auto_cdp_proc and auto_cdp_proc.poll() is None:
                            try:
                                auto_cdp_proc.terminate()
                            except Exception:
                                pass
                        auto_cdp_proc = None
                else:
                    print("⚠️ Chrome 실행 파일을 찾지 못해 자동 CDP를 건너뜁니다.")
            except Exception as _auto_cdp_e:
                print(f"⚠️ 자동 CDP 시도 실패({_auto_cdp_e}), smart 전략으로 진행합니다.")
                if auto_cdp_proc and auto_cdp_proc.poll() is None:
                    try:
                        auto_cdp_proc.terminate()
                    except Exception:
                        pass
                auto_cdp_proc = None
                
            # CDP 실패 시 일반 브라우저 실행
            if browser is None:
                launch_kwargs = dict(
                    user_data_dir=user_data_dir,
                    headless=False,
                    viewport={'width': 1920, 'height': 1080},
                    args=[
                        "--start-maximized",
                        "--window-size=1920,1080",
                        "--no-first-run",
                        "--no-default-browser-check",
                        "--disable-blink-features=AutomationControlled",
                        "--exclude-switches=enable-automation",
                    ]
                )
                if profile_strategy == 'system' and chrome_profile_dir_name:
                    launch_kwargs['args'].append(f"--profile-directory={chrome_profile_dir_name}")

                preferred_channel = os.getenv('CHROME_CHANNEL', 'chrome')
                try:
                    browser = await p.chromium.launch_persistent_context(
                        channel=preferred_channel,
                        **launch_kwargs,
                    )
                except Exception as e1:
                    print(f"⚠️ 첫 번째 실행(ch={preferred_channel}) 실패, 기본 Chromium으로 재시도: {e1}")
                    browser = await p.chromium.launch_persistent_context(**launch_kwargs)
            
            if 'page' not in locals():
                page = await browser.new_page()
                await page.set_viewport_size({'width': 1920, 'height': 1080})
            
            # 봇 감지 회피 스크립트
            await page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                window.chrome = { runtime: {}, loadTimes: function() {}, csi: function() {}, app: {} };
                Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
                Object.defineProperty(navigator, 'languages', { get: () => ['ko-KR', 'ko', 'en-US', 'en'] });
            """)
            
            log_progress("✨ 브라우저 생성 완료")
            print("🌐 네이버 쇼핑 접속 중...")
            await page.goto("https://shopping.naver.com/ns/home", wait_until='networkidle', timeout=30000)
            log_progress("✅ 네이버 쇼핑 페이지 로드 완료")
            print("💎 페이지 로드 완료! 원하는 카테고리나 검색을 수행한 후 Premium Crawler 버튼을 클릭하세요.")
            
            await wait_for_user_ready(page)
            log_progress("🎯 사용자 시작 신호 수신, 수집 시작")
            print("🚀 Premium Crawling 시작... (무제한 모드)")
            t0 = perf_counter()
            products_data = await scroll_and_collect(page)
            log_progress(f"✅ 수집 완료: {len(products_data)}개 ({perf_counter()-t0:.2f}s)")
            
            if products_data:
                print(f"💎 크롤링 완료! {len(products_data)}개 상품 수집")
                
                if highlight_keywords:
                    log_progress("🔍 키워드 매칭 시작")
                    matched_count = log_keyword_matching(products_data, highlight_keywords)
                    print(f"🎯 키워드 매칭된 상품: {matched_count}개")
                else:
                    log_keyword_matching(products_data, highlight_keywords)
                
                log_progress("💾 저장 단계 시작")
                html_filename = await save_products_data(products_data, highlight_keywords, page=page)
                log_progress("✅ 저장 단계 완료")
                
                if html_filename:
                    if os.getenv('SKIP_OPEN_HTML', '0') not in ('1', 'true', 'True'):
                        open_html_result(html_filename)
            else:
                print("⚠️ 크롤링 결과가 없습니다.")
            
            if os.getenv('AUTO_EXIT', '0') in ('1', 'true', 'True'):
                print("\n🚪 AUTO_EXIT 활성화: 엔터 대기 없이 종료합니다.")
            else:
                print("\n💎 결과 파일을 확인하시고 엔터를 누르면 종료됩니다.")
                input("🎯 엔터 키를 누르세요...")
            
        except Exception as e:
            # smart 전략: system 시도 중 실패했으면 즉시 local로 1회 더 재시도
            if profile_strategy == 'smart':
                print(f"⚠️ system 프로필 실행 실패로 smart 폴백(local) 시도: {e}")
                user_data_dir2 = get_or_create_local_chrome_profile()
                print(f"🔧 폴백 User Data: {user_data_dir2}")
                launch_kwargs = dict(
                    user_data_dir=user_data_dir2,
                    headless=False,
                    viewport={'width': 1920, 'height': 1080},
                    args=[
                        "--start-maximized",
                        "--window-size=1920,1080",
                        "--no-first-run",
                        "--no-default-browser-check",
                        "--disable-blink-features=AutomationControlled",
                        "--exclude-switches=enable-automation",
                    ]
                )
                preferred_channel = os.getenv('CHROME_CHANNEL', 'chrome')
                try:
                    browser = await p.chromium.launch_persistent_context(channel=preferred_channel, **launch_kwargs)
                except Exception:
                    browser = await p.chromium.launch_persistent_context(**launch_kwargs)
                page = await browser.new_page()
                await page.set_viewport_size({'width': 1920, 'height': 1080})
                print("⚠️ smart 폴백으로 브라우저는 실행되었지만, 재시작 흐름은 수동으로 재실행해 주세요.")
            else:
                print(f"❌ 오류 발생: {e}")
                import traceback
                traceback.print_exc()
                print("\n📋 문제 해결 가이드:")
                print("📍 시스템 Chrome 프로필 사용은 캡챠/로그인 유지에 유리하지만, 버전/락/정책으로 실패할 수 있습니다.")
                print("📍 PROFILE_STRATEGY=system|local|smart 로 선택 가능. 권장: smart")
                print("📍 CHROME_PROFILE_DIR=Default (또는 'Profile 1') 로 서브 프로필 지정 가능")
                print("📍 CHROME_CHANNEL=chrome|msedge 로 채널 전환 가능")
                print("📍 실패 시 모든 Chrome/Chromium 프로세스 종료 후 재시도 또는 'chrome-user-data' 삭제로 초기화")
        finally:
            try:
                if browser:
                    await browser.close()
                if auto_cdp_proc and auto_cdp_proc.poll() is None:
                    auto_cdp_proc.terminate()
            except Exception:
                pass

if __name__ == "__main__":
    print("\n" + "━" * 70)
    print("  💎 Premium Naver Shopping Intelligence System v7.0")
    print("━" * 70)
    print("")
    print("  ✨ Premium Features:")
    print("     ▸ Excel 빨간색 배경 셀 키워드 자동 추출")
    print("     ▸ 실시간 상품명 & 판매처 키워드 매칭")
    print("     ▸ 프리미엄 HTML 리포트 생성")
    print("     ▸ 인터랙티브 키워드 하이라이트 시스템")
    print("")
    print("━" * 70)
    print("  📊 Step 1: 엑셀 파일 선택 (키워드 추출)")
    print("  🚀 Step 2: 네이버 쇼핑 Premium Crawling")
    print("  💎 Step 3: Premium Analytics Report 생성")
    print("━" * 70)
    print("")
    asyncio.run(access_naver_shopping_optimized())